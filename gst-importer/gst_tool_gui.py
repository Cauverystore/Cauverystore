# -*- coding: utf-8 -*-
"""
GST Tool - a small visual interface for the GST master importer.

Everything goes through the importer API (URL + token below), which writes
into PostgreSQL - no database access is needed on this machine.

Buttons:
  Rebuild workbook from data  - recreate GST_Master.xlsx from the bundled
                                HSN/SAC/rate JSON (the five page sheets empty)
  Refresh workbook pages      - fetch the GSTN Master Codes page (needs an
                                Indian network) and fill the five page sheets
  Upload workbook to server   - push the workbook; the API imports it into
                                the database
  Download page -> upload     - fetch the page here, upload it directly to
                                the API (skip the workbook entirely)
  Run importer now            - the same call the web admin button makes
  Check last run              - last /update-gst-master result
  Open workbook in Excel      - handy for editing

Run:      python gst_tool_gui.py              (stdlib + openpyxl only)
Selftest: python gst_tool_gui.py --selftest   (no window, exits 0/1)
"""

import argparse
import io
import json
import os
import queue
import re
import sys
import threading
import time
import urllib.request
import webbrowser
from contextlib import redirect_stdout

import openpyxl

API_DEFAULT = "https://gst-importer-api-production.up.railway.app"
PAGE_URL = "https://einvoice1.gst.gov.in/Others/MasterCodes"
RETRIES = 3
TIMEOUT = 90

TAG_RE = re.compile(r"<[^>]+>")


def fetch_text(url, timeout=TIMEOUT, attempts=RETRIES):
    last = None
    for attempt in range(1, attempts + 1):
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                return resp.read().decode("utf-8", errors="replace")
        except Exception as exc:  # noqa: BLE001 - retried below
            last = exc
            if attempt < attempts:
                time.sleep(attempt * 5)
    raise last


def api_post(api_url, token, path, payload=None, timeout=1200):
    headers = {"Authorization": "Bearer " + token}
    if payload is not None:
        headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    req = urllib.request.Request(api_url.rstrip("/") + path, data=payload, headers=headers,
                                 method="GET" if payload is None else "POST")
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return resp.read().decode("utf-8", errors="replace")


def clean_cell(text):
    text = TAG_RE.sub("", text)
    for entity, char in (("&nbsp;", " "), ("&amp;", "&"), ("&lt;", "<"),
                         ("&gt;", ">"), ("&#39;", "'"), ("&quot;", '"')):
        text = text.replace(entity, char)
    return re.sub(r"\s+", " ", text).strip()


def parse_tables(html):
    out = []
    for table in re.findall(r"<table[^>]*>(.*?)</table>", html, re.S | re.I):
        rows = []
        for row in re.findall(r"<tr[^>]*>(.*?)</tr>", table, re.S | re.I):
            cells = [clean_cell(c) for c in re.findall(r"<t[dh][^>]*>(.*?)</t[dh]>", row, re.S | re.I)]
            if any(cells):
                rows.append(cells)
        if rows:
            out.append(rows)
    return out


def header_matches(header, expected):
    joined = " ".join(header).lower()
    return all(all(word in joined for word in pattern.split()) for pattern in expected)


PAGE_PATTERNS = [
    ("state_master", ["state code", "state name"]),
    ("country_master", ["country code", "country name"]),
    ("currency_master", ["currency code", "currency name"]),
    ("port_master", ["port code", "port name"]),
    ("uqc_master", ["unit code", "unit description"]),
]


def master_codes_from_html(html):
    found = {}
    for table in parse_tables(html):
        header = table[0] if table else []
        for name, patterns in PAGE_PATTERNS:
            if name in found or not header_matches(header, patterns):
                continue
            records = []
            for cells in table[1:]:
                if len(cells) < 2:
                    continue
                code, desc = cells[0], cells[1]
                if name == "state_master" and code.isdigit():
                    code = code.zfill(2)
                if code:
                    records.append((code, desc))
            found[name] = records
            print("OK    %-16s %5d rows" % (name, len(records)))
    for name, _patterns in PAGE_PATTERNS:
        if name not in found:
            print("WARN  table '%s' not found on the page - did it change?" % name)
    return found


def write_sheet(wb, title, header, rows):
    ws = wb.create_sheet(title)
    ws.append(header)
    for row in rows:
        ws.append(row)
    return ws


def runbook_rows():
    return [
        ["GST MASTER SHEET - edit the data sheets below, then use GST_Tool/GUI to upload."],
        ["1. Daily 02:00 IST the server syncs HSN/SAC/rates automatically."],
        ["2. Refresh the five code lists: GST Tool -> 'Refresh workbook pages'."],
        ["3. Upload: GST Tool -> 'Upload workbook to server'."],
        ["4. The web admin GST button makes the same API call."],
        ["5. Check a run: 'Check last run' in GST Tool, or SELECT * FROM gst_importer_import_logs."],
    ]


def rebuild_workbook(path, here):
    data = os.path.join(here, "data")
    with open(os.path.join(data, "hsn_master.json"), "r", encoding="utf-8") as fh:
        hsn_all = json.load(fh)
    with open(os.path.join(data, "gst_rate_seed.json"), "r", encoding="utf-8") as fh:
        seed = json.load(fh)
    hsn = [r for r in hsn_all if not str(r.get("code", "")).startswith("99")]
    sac = [r for r in hsn_all if str(r.get("code", "")).startswith("99")]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RUNBOOK"
    for row in runbook_rows():
        ws.append(row)
    ws.column_dimensions["A"].width = 110
    write_sheet(wb, "hsn", ["code", "description"],
                [[r.get("code"), r.get("description", "")] for r in hsn])
    write_sheet(wb, "sac", ["code", "description"],
                [[r.get("code"), r.get("description", "")] for r in sac])
    write_sheet(wb, "rates", ["code", "gst_rate", "effective_date"],
                [[r.get("hsnCode"), r.get("gstRate"), r.get("effectiveFrom")] for r in seed])
    for name in ("state_master", "country_master", "currency_master", "port_master", "uqc_master"):
        write_sheet(wb, name, ["code", "description"], [])
    wb.save(path)
    print("wrote %s (hsn=%d sac=%d rates=%d)" % (path, len(hsn), len(sac), len(seed)))


class GstToolApp:
    def __init__(self, root):
        self.root = root
        root.title("GST Tool - Noyyal Store")
        root.geometry("880x660")
        self.queue = queue.Queue()
        self.busy = False
        self.config = self._load_config()
        self._build_ui()
        self._console("GST Tool ready. API: %s" % self.entry_api.get().strip())
        self.root.after(120, self._poll_queue)

    # ------------------------------------------------------------- config
    def _config_path(self):
        base = os.environ.get("LOCALAPPDATA", os.path.expanduser("~"))
        folder = os.path.join(base, "NoyyalStore")
        os.makedirs(folder, exist_ok=True)
        return os.path.join(folder, "gst_tool.json")

    def _load_config(self):
        try:
            with open(self._config_path(), "r", encoding="utf-8") as fh:
                return json.load(fh)
        except Exception:  # noqa: BLE001 - first run
            return {}

    def _save_config(self):
        try:
            with open(self._config_path(), "w", encoding="utf-8") as fh:
                json.dump({"api_url": self.entry_api.get().strip(),
                           "token": self.entry_token.get().strip(),
                           "workbook": self.entry_book.get().strip()}, fh, indent=2)
        except Exception:  # noqa: BLE001 - non-fatal
            pass

    def _default_workbook(self):
        here = os.path.dirname(os.path.abspath(__file__))
        return os.path.join(here, "GST_Master.xlsx")

    # ------------------------------------------------------------------ UI
    def _build_ui(self):
        import tkinter as tk
        from tkinter import filedialog, messagebox, scrolledtext, ttk

        self._tk = tk
        self._filedialog = filedialog
        self._messagebox = messagebox

        frame = ttk.Frame(self.root, padding=10)
        frame.pack(fill="both", expand=True)

        server = ttk.LabelFrame(frame, text="Server (importer API)", padding=8)
        server.pack(fill="x", pady=4)
        ttk.Label(server, text="API URL").grid(row=0, column=0, sticky="w")
        self.entry_api = ttk.Entry(server, width=54)
        self.entry_api.grid(row=0, column=1, sticky="we", padx=6)
        self.entry_api.insert(0, self.config.get("api_url", API_DEFAULT))
        ttk.Label(server, text="Token").grid(row=1, column=0, sticky="w")
        self.entry_token = ttk.Entry(server, width=54, show="*")
        self.entry_token.grid(row=1, column=1, sticky="we", padx=6)
        self.entry_token.insert(0, self.config.get("token", ""))
        ttk.Button(server, text="Test connection", command=self.task_test).grid(
            row=0, column=2, rowspan=2, sticky="ns", padx=8)
        server.columnconfigure(1, weight=1)

        work = ttk.LabelFrame(frame, text="Workbook", padding=8)
        work.pack(fill="x", pady=4)
        ttk.Label(work, text="GST_Master.xlsx").grid(row=0, column=0, sticky="w")
        self.entry_book = ttk.Entry(work, width=62)
        self.entry_book.grid(row=0, column=1, sticky="we", padx=6)
        self.entry_book.insert(0, self.config.get("workbook", self._default_workbook()))
        ttk.Button(work, text="Browse...", command=self._browse_book).grid(row=0, column=2, padx=8)
        work.columnconfigure(1, weight=1)

        actions = ttk.LabelFrame(frame, text="Actions", padding=8)
        actions.pack(fill="x", pady=4)
        self.buttons = []
        specs = [
            ("Rebuild workbook from data", self.task_rebuild),
            ("Refresh workbook pages", self.task_refresh_pages),
            ("Upload workbook to server", self.task_upload_book),
            ("Download page and upload", self.task_page_to_server),
            ("Run importer now", self.task_run_importer),
            ("Check last run", self.task_status),
            ("Open workbook in Excel", self.task_open_book),
        ]
        for i, (text, cmd) in enumerate(specs):
            b = ttk.Button(actions, text=text, command=cmd)
            b.grid(row=i // 3, column=i % 3, sticky="we", padx=4, pady=3)
            self.buttons.append(b)
        for i in range(3):
            actions.columnconfigure(i, weight=1)

        console_frame = ttk.LabelFrame(frame, text="Console", padding=8)
        console_frame.pack(fill="both", expand=True, pady=4)
        self.console = scrolledtext.ScrolledText(console_frame, height=13, wrap="word",
                                                 font=("Consolas", 9))
        self.console.pack(fill="both", expand=True)
        self.console.config(state="disabled")

        self.status = ttk.Label(frame, text="Idle", relief="sunken", anchor="w")
        self.status.pack(fill="x")

    # ------------------------------------------------------------------ util
    def _browse_book(self):
        path = self._filedialog.askopenfilename(
            title="Choose the GST master workbook",
            filetypes=[("Excel workbook", "*.xlsx")])
        if path:
            self.entry_book.delete(0, "end")
            self.entry_book.insert(0, path)

    def _console(self, text):
        self.queue.put(("line", text))

    def _status_text(self, text):
        self.queue.put(("status", text))

    def _busy(self, on):
        self.busy = on
        state = "disabled" if on else "normal"
        for button in self.buttons:
            button.configure(state=state)

    def _globals(self):
        self._save_config()

    def _poll_queue(self):
        try:
            while True:
                kind, payload = self.queue.get_nowait()
                if kind == "line":
                    self.console.config(state="normal")
                    self.console.insert("end", payload + "\n")
                    self.console.see("end")
                    self.console.config(state="disabled")
                elif kind == "status":
                    self.status.config(text=payload)
                elif kind == "done":
                    self.console.config(state="normal")
                    self.console.insert("end", payload + "\n")
                    self.console.see("end")
                    self.console.config(state="disabled")
                    self._busy(False)
                    self.status.config(text="Ready")
        except queue.Empty:
            pass
        self.root.after(100, self._poll_queue)

    def _run(self, fn, label):
        if self.busy:
            return
        self._save_config()
        self._busy(True)
        self._status_text("running: %s" % label)
        threading.Thread(target=self._worker, args=(fn,), daemon=True).start()

    def _worker(self, fn):
        buffer = io.StringIO()
        try:
            with redirect_stdout(buffer):
                try:
                    fn()
                except Exception as exc:  # noqa: BLE001 - shown in the console
                    print("ERROR %s: %s" % (type(exc).__name__, exc))
        finally:
            self.queue.put(("done", buffer.getvalue()))

    def _creds(self):
        if not self.entry_api.get().strip():
            raise RuntimeError("API URL is empty")
        if not self.entry_token.get().strip():
            raise RuntimeError("Token is empty")

    # ------------------------------------------------------------------ tasks
    def task_test(self):
        self._run(self._test, "test connection")

    def _test(self):
        self._creds()
        print("GET %s/ " % self.entry_api.get().strip())
        print(api_post(self.entry_api.get(), self.entry_token.get(), "/"))

    def task_rebuild(self):
        self._run(self._rebuild, "rebuild workbook")

    def _rebuild(self):
        path = self.entry_book.get().strip() or self._default_workbook()
        here = os.path.dirname(os.path.abspath(__file__))
        rebuild_workbook(path, here)

    def task_refresh_pages(self):
        self._run(self._refresh_pages, "refresh workbook pages")

    def _refresh_pages(self):
        path = self.entry_book.get().strip()
        if not os.path.exists(path):
            raise RuntimeError("Workbook not found: %s" % path)
        print("downloading %s ..." % PAGE_URL)
        html = fetch_text(PAGE_URL)
        page = master_codes_from_html(html)
        wb = openpyxl.load_workbook(path)
        for name, _patterns in PAGE_PATTERNS:
            if name not in wb.sheetnames:
                print("WARN  sheet '%s' missing - skip (rebuild the workbook)" % name)
                continue
            ws = wb[name]
            ws.delete_rows(1, ws.max_row)
            ws.append(["code", "description"])
            for code, desc in page.get(name, []):
                ws.append([code, desc])
        wb.save(path)
        print("workbook page tables updated - now upload it ('Upload workbook to server')")

    def task_upload_book(self):
        self._run(self._upload_book, "upload workbook")

    def _upload_book(self):
        self._creds()
        path = self.entry_book.get().strip()
        with open(path, "rb") as fh:
            payload = fh.read()
        print("uploading %s (%d bytes) ..." % (path, len(payload)))
        print(api_post(self.entry_api.get(), self.entry_token.get(), "/upload-master-sheet", payload))

    def task_page_to_server(self):
        self._run(self._page_to_server, "download page")

    def _page_to_server(self):
        self._creds()
        print("downloading %s ..." % PAGE_URL)
        html = fetch_text(PAGE_URL)
        print("uploading page (%d chars) ..." % len(html))
        print(api_post(self.entry_api.get(), self.entry_token.get(),
                       "/upload-master-codes", html.encode("utf-8")))

    def task_run_importer(self):
        self._run(self._run_importer, "run importer")

    def _run_importer(self):
        self._creds()
        print("POST /update-gst-master (the web admin button does exactly this)")
        print(api_post(self.entry_api.get(), self.entry_token.get(), "/update-gst-master"))

    def task_status(self):
        self._run(self._status, "check last run")

    def _status(self):
        self._creds()
        print(api_post(self.entry_api.get(), self.entry_token.get(), "/update-gst-master/status"))

    def task_open_book(self):
        path = self.entry_book.get().strip()
        if not os.path.exists(path):
            self._messagebox.showerror("GST Tool", "Workbook not found:\n%s" % path)
            return
        webbrowser.open(path)


def selftest():
    sample = """<html><body>
      <table><tr><th>State Code</th><th>State Name</th></tr>
      <tr><td>1</td><td>Jammu and Kashmir</td></tr></table>
      <table><tr><th>Country Code</th><th>Country Name</th></tr>
      <tr><td>IN</td><td>India</td></tr></table>
      </body></html>"""
    page = master_codes_from_html(sample)
    assert page["state_master"] == [("01", "Jammu and Kashmir")], page["state_master"]
    assert page["country_master"] == [("IN", "India")], page["country_master"]
    wb = openpyxl.Workbook()
    write_sheet(wb, "hsn", ["code", "description"], [["0101", "desc"]])
    assert wb["hsn"]["A2"].value == "0101"
    print("selftest OK")
    return 0


def main():
    parser = argparse.ArgumentParser(description="GST Tool - visual interface")
    parser.add_argument("--selftest", action="store_true", help="run checks and exit")
    args = parser.parse_args()
    if args.selftest:
        return selftest()
    import tkinter as tk
    root = tk.Tk()
    GstToolApp(root)
    root.mainloop()
    return 0


if __name__ == "__main__":
    sys.exit(main())