# -*- coding: utf-8 -*-
"""
Builds GST_Master.xlsx - the editable master workbook for the whole importer -
from the committed official data (HSN / SAC lists, CBIC rate seed) plus a live
pull of the GSTN Master Codes page for the five code lists.

Usage:
    python build_master_sheet.py [--no-page]

Sheets:
    RUNBOOK  how the tool is operated day to day (edit the data sheets below)
    hsn      code | description                  from data/hsn_master.json
    sac      code | description                  from data/sac_master.json
    rates    code | gst_rate | effective_date    from data/gst_rate_seed.json
    state_master / country_master / currency_master / port_master / uqc_master
             code | description                  live page (skipped with --no-page)
"""

import argparse
import json
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.join(HERE, "data")
OUT = os.path.join(HERE, "GST_Master.xlsx")

INTRO = [
    ["GST MASTER SHEET - Noyyal Store"],
    ["How to run the tool"],
    ["1. Daily 02:00 IST  - automatic (Railway cron '30 20 * * *' UTC); the cron syncs HSN/SAC/rates from "
     "the data bundled with the container (no network needed)."],
    ["2. Monthly  - refresh the five code lists: run  python fetch_upload_master.py  on this machine "
     "(GSTN only answers reliably from India; it uploads the page to the API and the API writes the tables)."],
    ["3. On any rate change  - edit the hsn / sac / rates sheets below, save, then run "
     "python push_master_sheet.py (uploads this workbook and imports it into the database)."],
    ["4. From the web app - Admin > GST Rates: the refresh button tries the importer API first and falls back "
     "to the store's own refresh."],
    ["5. Verify - GET /update-gst-master/status on the importer API, or SELECT * FROM gst_importer_import_logs "
     "in Postgres, newest first."],
    ["6. Windows exe - rebuilt by GitHub Actions on every push (download from Actions > artifacts)."],
    ["", "Only the sheets below are data - do not rename columns or sheets."],
]

PAGE_SHEETS = ["state_master", "country_master", "currency_master", "port_master", "uqc_master"]


def read_json(name):
    with open(os.path.join(DATA, name), "r", encoding="utf-8") as fh:
        return json.load(fh)


def write_sheet(wb, title, header, rows, widths):
    ws = wb.create_sheet(title)
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E7EFE7")
    for row in rows:
        ws.append(row)
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    return ws


def main():
    parser = argparse.ArgumentParser(description="Build GST_Master.xlsx from the committed data")
    parser.add_argument("--no-page", action="store_true", help="skip the live Master Codes fetch")
    args = parser.parse_args()

    wb = Workbook()
    ws = wb.active
    ws.title = "RUNBOOK"
    for row in INTRO:
        ws.append(row)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for i, width in enumerate((120,), start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    hsn_all = read_json("hsn_master.json")
    seed = read_json("gst_rate_seed.json")
    hsn = [r for r in hsn_all if not str(r.get("code", "")).startswith("99")]
    sac = [r for r in hsn_all if str(r.get("code", "")).startswith("99")]
    write_sheet(wb, "hsn", ["code", "description"],
                [[r.get("code"), r.get("description", "")] for r in hsn], (18, 90))
    write_sheet(wb, "sac", ["code", "description"],
                [[r.get("code"), r.get("description", "")] for r in sac], (18, 90))
    write_sheet(wb, "rates", ["code", "gst_rate", "effective_date"],
                [[r.get("hsnCode"), r.get("gstRate"), r.get("effectiveFrom")] for r in seed],
                (18, 12, 16))

    page = {}
    if not args.no_page:
        print("fetching the Master Codes page (needs Indian network) ...")
        import gst_master_importer as imp
        page = imp.fetch_master_codes()
    for name in PAGE_SHEETS:
        write_sheet(wb, name, ["code", "description"], page.get(name, []), (18, 60))

    wb.save(OUT)
    print("wrote %s" % OUT)
    print("hsn=%d sac=%d rates=%d page=%s" % (len(hsn), len(sac), len(seed),
                                              {k: len(v) for k, v in page.items()}))


if __name__ == "__main__":
    sys.exit(main())