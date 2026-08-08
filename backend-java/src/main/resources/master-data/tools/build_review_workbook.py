# -*- coding: utf-8 -*-
"""
Builds the spreadsheet a reviewer works through to approve GST rates.

Every rate the seed could not approve on its own lands here. They are not wrong - they are
undecided: the heading carries more than one published rate and only the goods description
says which applies. Nothing unapproved is ever charged, so a product under one of these
headings cannot be invoiced at all until someone signs it off.

The rows are grouped by HSN code with a "competing rates" column, because a reviewer cannot
choose between rates they cannot see side by side. Only the yellow columns are theirs to fill.

Re-run whenever the seed changes; the counts on the Summary sheet are live formulas.

    python src/main/resources/master-data/tools/build_review_workbook.py
"""

import collections
import json
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

DATA = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..")
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "GST_unverified_headings_for_review.xlsx")

ARIAL = "Arial"
HEAD_FILL = PatternFill("solid", fgColor="1F3864")
INPUT_FILL = PatternFill("solid", fgColor="FFFF00")
BAND_FILL = PatternFill("solid", fgColor="F2F2F2")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

with open(os.path.join(DATA, "gst_rate_seed.json"), encoding="utf-8") as fh:
    seed = json.load(fh)
with open(os.path.join(DATA, "hsn_master.json"), encoding="utf-8") as fh:
    hsn = json.load(fh)
with open(os.path.join(DATA, "chapter_master.json"), encoding="utf-8") as fh:
    chapters = {c["chapter"]: c["title"] for c in json.load(fh)}

by_code = collections.defaultdict(list)
for row in seed:
    by_code[row["hsnCode"]].append(row)

hsn_desc = {h["code"]: h["description"].split("~")[-1] for h in hsn}

# One row per unverified rate, grouped so competing rates on a heading sit together - a
# reviewer cannot choose between rates they cannot see side by side.
records = []
for code in sorted(by_code):
    rows = by_code[code]
    if not any(r["status"] != "VERIFIED" for r in rows):
        continue
    competing = " / ".join("%g%%" % r for r in sorted({x["gstRate"] for x in rows}))
    for r in sorted(rows, key=lambda x: x["gstRate"]):
        if r["status"] == "VERIFIED":
            continue
        cond = r.get("conditionType") or "NONE"
        if cond == "VALUE_UPTO":
            cond_text = "Up to Rs %g per %s" % (r.get("thresholdAmount") or 0, r.get("thresholdUnit") or "piece")
        elif cond == "VALUE_ABOVE":
            cond_text = "Above Rs %g per %s" % (r.get("thresholdAmount") or 0, r.get("thresholdUnit") or "piece")
        elif cond == "PRE_PACKAGED":
            cond_text = "Pre-packaged and labelled"
        elif cond == "NOT_PRE_PACKAGED":
            cond_text = "Sold loose"
        else:
            cond_text = "-"
        records.append({
            "code": code,
            "level": {2: "Chapter", 4: "Heading", 6: "Sub-heading", 8: "Tariff item"}.get(len(code), "%d-digit" % len(code)),
            "chapter": chapters.get(code[:2]) or "",
            "goods": hsn_desc.get(code, "")[:180],
            "rate": r["gstRate"] / 100.0,
            "competing": competing,
            "condition": cond_text,
            "cbic": (r.get("conditionText") or "").strip()[:400],
            "source": (r.get("source") or "").strip()[:160],
        })

wb = Workbook()

# ---------------------------------------------------------------- how to use
gui = wb.active
gui.title = "How to use"
gui.sheet_view.showGridLines = False
guide = [
    ("GST rate review queue", 16, True),
    ("", 11, False),
    ("What this is", 12, True),
    ("Every GST rate the Cauvery Store seed could not approve on its own. They are not wrong -", 11, False),
    ("they are undecided. Each heading below carries more than one published rate, and only the", 11, False),
    ("goods description says which applies. Until a rate is approved it is never charged, so any", 11, False),
    ("product under these headings cannot be invoiced at all.", 11, False),
    ("", 11, False),
    ("What to do", 12, True),
    ("Work down the 'Review queue' sheet. Rows for the same HSN code sit together, and the", 11, False),
    ("'Competing rates' column shows every rate published for that heading, so the choice is", 11, False),
    ("visible in one place. Read the CBIC wording, decide whether it describes goods this store", 11, False),
    ("sells, and fill in the four yellow columns.", 11, False),
    ("", 11, False),
    ("Only the yellow columns are for you: Decision, Approved rate, Reviewed by, Date, Notes.", 11, True),
    ("Everything to their left is the published data and should not be edited.", 11, False),
    ("", 11, False),
    ("Decision may be APPROVE, REJECT or HOLD.", 11, False),
    ("   APPROVE - this rate is correct for goods sold under this code; it becomes chargeable.", 11, False),
    ("   REJECT  - this rate does not apply to anything this store sells; leave it unused.", 11, False),
    ("   HOLD    - cannot decide without more information. Say what is needed in Notes.", 11, False),
    ("", 11, False),
    ("Row 2 of the review queue is a filled-in example showing the expected format.", 11, True),
    ("Delete it before returning the file.", 11, False),
    ("", 11, False),
    ("Where the rates come from", 12, True),
    ("CBIC Notification 09/2025-Central Tax (Rate) as amended by 19/2025 (eff. 01-02-2026) and", 11, False),
    ("01/2026 (eff. 01-05-2026). The 'Source' column names the notification for each row.", 11, False),
    ("Decide against the notification itself, not against a summary published elsewhere -", 11, False),
    ("summaries routinely flatten the price bands and packaging splits that make these ambiguous.", 11, False),
]
for i, (text, size, bold) in enumerate(guide, start=1):
    c = gui.cell(row=i, column=1, value=text)
    c.font = Font(name=ARIAL, size=size, bold=bold)
gui.column_dimensions["A"].width = 100

# ---------------------------------------------------------------- review queue
ws = wb.create_sheet("Review queue")
headers = ["HSN / Chapter", "Level", "Chapter", "Goods (official master)", "Rate",
           "Competing rates", "Condition", "CBIC wording for this rate", "Source",
           "Decision", "Approved rate", "Reviewed by", "Date (dd-mm-yyyy)", "Notes"]
widths = [14, 12, 30, 44, 9, 16, 22, 62, 34, 13, 13, 20, 17, 34]
for col, (name, width) in enumerate(zip(headers, widths), start=1):
    c = ws.cell(row=1, column=col, value=name)
    c.font = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
    c.fill = HEAD_FILL
    c.alignment = Alignment(vertical="center", wrap_text=True)
    c.border = BOX
    ws.column_dimensions[get_column_letter(col)].width = width
ws.row_dimensions[1].height = 30

INPUT_COLS = (10, 11, 12, 13, 14)

def write_row(r, rec, example=False):
    values = [rec["code"], rec["level"], rec["chapter"], rec["goods"], rec["rate"],
              rec["competing"], rec["condition"], rec["cbic"], rec["source"]]
    for col, v in enumerate(values, start=1):
        c = ws.cell(row=r, column=col, value=v)
        c.font = Font(name=ARIAL, size=10)
        c.alignment = Alignment(vertical="top", wrap_text=col in (3, 4, 7, 8, 9))
        c.border = BOX
        if col == 5:
            c.number_format = "0.00%"
    for col in INPUT_COLS:
        c = ws.cell(row=r, column=col)
        c.fill = INPUT_FILL
        c.font = Font(name=ARIAL, size=10)
        c.border = BOX
        c.alignment = Alignment(vertical="top", wrap_text=(col == 14))
        if col == 11:
            c.number_format = "0.00%"

# A worked example so the expected format is unambiguous. Deleted before the file is used.
write_row(2, {
    "code": "1905", "level": "Heading", "chapter": "PREPARATIONS OF CEREALS, FLOUR, STARCH OR MILK",
    "goods": "BREAD, PASTRY, CAKES, BISCUITS AND OTHER BAKERS' WARES",
    "rate": 0.05, "competing": "0% / 5%", "condition": "-",
    "cbic": "Pastry, cakes, biscuits and other bakers' wares, whether or not containing cocoa",
    "source": "CBIC Notif. 09/2025-CT(Rate)",
}, example=True)
ws.cell(row=2, column=10, value="APPROVE")
ws.cell(row=2, column=11, value=0.05).number_format = "0.00%"
ws.cell(row=2, column=12, value="EXAMPLE ROW - delete me")
ws.cell(row=2, column=13, value="08-08-2026")
ws.cell(row=2, column=14, value="We sell biscuits and cakes; the nil row is for pappad, which we do not sell.")
for col in range(1, 15):
    ws.cell(row=2, column=col).font = Font(name=ARIAL, size=10, italic=True, color="808080")

start = 3
prev_code = None
shade = False
for i, rec in enumerate(records):
    r = start + i
    if rec["code"] != prev_code:
        shade = not shade
        prev_code = rec["code"]
    write_row(r, rec)
    if shade:
        for col in range(1, 10):
            ws.cell(row=r, column=col).fill = BAND_FILL

last = start + len(records) - 1
dv = DataValidation(type="list", formula1='"APPROVE,REJECT,HOLD"', allow_blank=True, showDropDown=False)
ws.add_data_validation(dv)
dv.add("J%d:J%d" % (start, last))
ws.freeze_panes = "A2"
ws.auto_filter.ref = "A1:N%d" % last

# ---------------------------------------------------------------- summary
sm = wb.create_sheet("Summary", 1)
sm.sheet_view.showGridLines = False
sm["A1"] = "Review progress"
sm["A1"].font = Font(name=ARIAL, size=14, bold=True)
rows = [
    ("Rates awaiting review", '=COUNTA(\'Review queue\'!A%d:A%d)' % (start, last)),
    ("Distinct HSN codes affected", None),
    ("", None),
    ("Approved", '=COUNTIF(\'Review queue\'!J%d:J%d,"APPROVE")' % (start, last)),
    ("Rejected", '=COUNTIF(\'Review queue\'!J%d:J%d,"REJECT")' % (start, last)),
    ("On hold", '=COUNTIF(\'Review queue\'!J%d:J%d,"HOLD")' % (start, last)),
    ("Still undecided", '=B3-B6-B7-B8'),
    ("", None),
    ("Percent complete", '=IFERROR((B6+B7+B8)/B3,0)'),
]
sm["A3"] = "Rates awaiting review"
r = 3
for label, formula in rows:
    sm.cell(row=r, column=1, value=label).font = Font(name=ARIAL, size=11, bold=label in ("Percent complete",))
    if formula:
        c = sm.cell(row=r, column=2, value=formula)
        c.font = Font(name=ARIAL, size=11)
        c.number_format = "0.0%" if label == "Percent complete" else "#,##0"
    r += 1
sm["B4"] = len({rec["code"] for rec in records})
sm["B4"].font = Font(name=ARIAL, size=11)
sm["B4"].number_format = "#,##0"
sm.column_dimensions["A"].width = 30
sm.column_dimensions["B"].width = 14

sm["A14"] = "Counts update themselves as the Decision column is filled in."
sm["A14"].font = Font(name=ARIAL, size=10, italic=True, color="808080")
sm["A15"] = "Source: Cauvery Store gst_rate_seed.json, derived from CBIC Notification 09/2025-CT(Rate) as amended."
sm["A15"].font = Font(name=ARIAL, size=10, italic=True, color="808080")
sm["A16"] = "Generated 08-08-2026. The counts start at row 3, so the example row is excluded from them."
sm["A16"].font = Font(name=ARIAL, size=10, italic=True, color="808080")

wb.save(OUT)
print("wrote %s" % OUT)
print("%d rate rows across %d HSN codes" % (len(records), len({r['code'] for r in records})))
